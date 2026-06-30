# utils/excel_exporter.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
from datetime import datetime
import tempfile
import os

class ExcelExporter:
    """Export scholarship results to Excel with formatting"""
    
    @staticmethod
    def export_scholarships(scholarships: List[Dict], filename: str = None) -> str:
        """
        Export scholarships to Excel file
        
        Args:
            scholarships: List of scholarship dictionaries
            filename: Output filename (auto-generated if None)
        
        Returns:
            Path to generated Excel file
        """
        # Ensure we always have a valid filename
        if not filename or not filename.strip():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"scholarships_{timestamp}"
        
        # Strip any path components and whitespace
        filename = os.path.basename(filename.strip())
        
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename = filename + '.xlsx'
        
        # Use temp directory for file creation (works on PythonAnywhere)
        filepath = os.path.join(tempfile.gettempdir(), filename)
        
        # Prepare data for DataFrame
        data = []
        for idx, sch in enumerate(scholarships, 1):
            data.append({
                'No.': idx,
                'Scholarship Title': sch.get('title', ''),
                'Country': sch.get('country', ''),
                'Degree Level': sch.get('degree', ''),
                'Field of Study': sch.get('field', ''),
                'Duration': sch.get('duration', ''),
                'Funding Coverage': sch.get('funding', ''),
                'Eligibility': sch.get('eligibility', ''),
                'Required Documents': sch.get('documents', ''),
                'Deadline': sch.get('deadline', ''),
                'Official Link': sch.get('url', ''),
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Write to Excel (explicitly set engine to avoid 'No engine for filetype' error)
        df.to_excel(filepath, index=False, sheet_name='Scholarships', engine='openpyxl')
        
        # Apply formatting
        ExcelExporter._apply_formatting(filepath)
        
        return filepath
    
    @staticmethod
    def _apply_formatting(filename: str):
        """Apply professional formatting to Excel file"""
        wb = load_workbook(filename)
        ws = wb.active
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border
        
        # Adjust column widths
        column_widths = {
            'A': 5,   # No.
            'B': 40,  # Title
            'C': 15,  # Country
            'D': 15,  # Degree
            'E': 20,  # Field
            'F': 12,  # Duration
            'G': 20,  # Funding
            'H': 35,  # Eligibility
            'I': 30,  # Documents
            'J': 15,  # Deadline
            'K': 50,  # Link
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        
        # Save formatted workbook
        wb.save(filename)